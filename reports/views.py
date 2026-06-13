import logging
from functools import wraps
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.core.exceptions import PermissionDenied
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.timezone import now

from appointments.models import Appointment
from specialists.models import Specialist

logger = logging.getLogger(__name__)


def owner_required(view_func):
    """Доступ только для суперпользователя или группы «Владелец»."""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            from django.contrib.auth.views import redirect_to_login
            return redirect_to_login(request.get_full_path())
        if not (
            request.user.is_superuser
            or request.user.groups.filter(name='Владелец').exists()
        ):
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return wrapper


@owner_required
def income_report(request):
    specialists          = Specialist.objects.filter(is_active=True)
    selected_specialist  = request.GET.get('specialist')
    date_from            = request.GET.get('date_from')
    date_to              = request.GET.get('date_to')

    appointments = (
        Appointment.objects
        .filter(status=Appointment.STATUS_COMPLETED)
        .select_related('client', 'specialist', 'service')
        .order_by('date', 'time_start')
    )

    if selected_specialist:
        appointments = appointments.filter(specialist_id=selected_specialist)
    if date_from:
        appointments = appointments.filter(date__gte=date_from)
    if date_to:
        appointments = appointments.filter(date__lte=date_to)

    # Итог через БД — один запрос вместо Python-цикла
    total = appointments.aggregate(total=Sum('service__price'))['total'] or 0

    if 'download' in request.GET:
        logger.info(
            'Пользователь %s скачал отчёт по доходам (специалист=%s, %s — %s)',
            request.user, selected_specialist, date_from, date_to,
        )
        response = _build_excel(appointments, total)
        return response

    return render(request, 'reports/income.html', {
        'specialists':         specialists,
        'appointments':        appointments,
        'selected_specialist': selected_specialist,
        'date_from':           date_from,
        'date_to':             date_to,
        'total':               total,
    })


def _build_excel(appointments, total):
    """Формирует Excel-файл отчёта и возвращает HttpResponse."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Отчёт по доходу'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='4F81BD')
    center      = Alignment(horizontal='center')

    headers = ['Дата', 'Специалист', 'Клиент', 'Услуга', 'Цена (руб.)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center

    for row, appt in enumerate(appointments, 2):
        ws.cell(row=row, column=1, value=str(appt.date))
        ws.cell(row=row, column=2, value=appt.specialist.full_name)
        ws.cell(row=row, column=3, value=appt.client.full_name)
        ws.cell(row=row, column=4, value=appt.service.name)
        ws.cell(row=row, column=5, value=float(appt.service.price))

    total_row = appointments.count() + 2
    ws.cell(row=total_row, column=4, value='Итого:').font = Font(bold=True)
    ws.cell(row=total_row, column=5, value=float(total)).font = Font(bold=True)

    for col, width in zip('ABCDE', [12, 25, 25, 30, 15]):
        ws.column_dimensions[col].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'отчёт_{now().strftime("%Y-%m-%d")}.xlsx'
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response